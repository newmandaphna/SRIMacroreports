"""Reading an uploaded workbook file, for importing historical data.

The database is the system of record and the sheets are only ingestion, so history
that predates the current quarterly sheet needs a way in. This is that way: an admin
uploads an .xlsx or .csv export and it flows through exactly the same pipeline as a
Google Sheets sync. Same column allowlist, same validation, same alias resolution,
same rejection queue, same audit trail. The file is parsed in memory and never
written to disk, so nothing PHI bearing lands in the container's filesystem.
"""

from __future__ import annotations

import csv
import io

from app.sync.sheets import (
    MAX_ROWS,
    SheetData,
    SheetsError,
    assert_not_truncated,
    assert_tab_allowed,
    build_sheet_data,
    selectable_tabs,
)

# A quarter of the real workbook is under 3 MB as xlsx. Ten times that is headroom,
# not an invitation.
MAX_UPLOAD_BYTES = 25 * 1024 * 1024

CSV_TAB_NAME = "Uploaded CSV"


class UploadedWorkbookClient:
    """SheetsClient over one uploaded file instead of a live spreadsheet.

    A CSV is a single table with no tab concept, so it answers any allowed tab name.
    An XLSX keeps its own tab names, and the RAW_ tab block applies to them exactly
    as it does on the live sheet.
    """

    def __init__(self, filename: str, content: bytes) -> None:
        if len(content) > MAX_UPLOAD_BYTES:
            raise SheetsError(
                f"That file is larger than {MAX_UPLOAD_BYTES // (1024 * 1024)} MB. "
                "Export the quarter you need rather than the whole history at once."
            )
        if not content:
            raise SheetsError("The uploaded file is empty.")

        name = (filename or "").lower()
        if name.endswith(".csv"):
            self._tabs = {CSV_TAB_NAME: _parse_csv(content)}
            self._single_table = True
        elif name.endswith(".xlsx"):
            self._tabs = _parse_xlsx(content)
            self._single_table = False
        else:
            raise SheetsError(
                "Only .xlsx and .csv files can be imported. Export the tab you need "
                "from the workbook as one of those."
            )

        if not self._tabs:
            raise SheetsError("The workbook has no readable tabs.")

    def list_tabs(self, spreadsheet_id: str) -> list[str]:
        return list(self._tabs)

    def read_tab(self, spreadsheet_id: str, tab_name: str, header_row: int) -> SheetData:
        resolved = self._resolve_tab(tab_name)
        assert_tab_allowed(resolved)
        return build_sheet_data(header_row, self._tabs[resolved])

    def _resolve_tab(self, tab_name: str | None) -> str:
        if self._single_table:
            # A CSV has exactly one table, whatever the source's tab field says.
            if tab_name:
                assert_tab_allowed(tab_name)
            return CSV_TAB_NAME
        if tab_name and tab_name in self._tabs:
            return tab_name
        allowed = selectable_tabs(list(self._tabs))
        if not tab_name and len(allowed) == 1:
            return allowed[0]
        raise SheetsError(
            f"The uploaded workbook has no tab named {tab_name!r}. "
            f"It contains: {', '.join(allowed) or 'no importable tabs'}. "
            "Set the source's Tab field to the one to import, or re-export the file."
        )


def _parse_csv(content: bytes) -> list[list[object]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        # Excel on Windows exports CSV as cp1252 unless told otherwise.
        text = content.decode("cp1252", errors="replace")
    reader = csv.reader(io.StringIO(text))
    # One past the cap, so hitting it is detectable rather than a silent trim.
    rows = [list(row) for _, row in zip(range(MAX_ROWS + 1), reader, strict=False)]
    assert_not_truncated(len(rows), CSV_TAB_NAME)
    return rows


def _parse_xlsx(content: bytes) -> dict[str, list[list[object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - a deployment problem, not logic
        raise SheetsError("The openpyxl library is not installed.") from exc

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise SheetsError(
            "That file could not be read as an Excel workbook. If it came from "
            "Google Sheets, use File, Download, Microsoft Excel (.xlsx)."
        ) from exc

    tabs: dict[str, list[list[object]]] = {}
    try:
        for sheet in workbook.worksheets:
            rows: list[list[object]] = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(list(row))
                if len(rows) > MAX_ROWS:
                    break
            assert_not_truncated(len(rows), sheet.title)
            tabs[sheet.title] = rows
    finally:
        workbook.close()
    return tabs
