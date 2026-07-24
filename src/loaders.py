"""Valant raw-export loaders (Phase A).

Valant reports carry quirks the loader must absorb rather than have anyone
hand-clean the file (which would hide the very problems the reconciliation
harness exists to catch). The "NewFeeSchedule" charges variant has a three-line
preamble (a Textbox row, a date row, a blank) before the real header; the header
row is detected **programmatically**, never `skiprows=3`.

No pandas: money columns must stay Decimal (ASSUMPTIONS §10), and pandas would
coerce them to float64 on read -- the exact silent-float bug the spec forbids.
CSV is read with the stdlib; XLSX support is optional and loaded lazily via
openpyxl.
"""
from __future__ import annotations

import csv
import os
from typing import Dict, List, Optional, Sequence, Tuple


class LoaderError(Exception):
    """Base class for loader failures."""


class HeaderNotFoundError(LoaderError):
    """No row matched enough expected header tokens (preamble detection failed)."""


class MissingColumnsError(LoaderError):
    """A required column is absent, blank, or duplicated after header detection."""


# Report family metadata. Canonical column names align with the compensation
# engine (newmandaphna/SRIcompensation) so the two projects call a column the
# same thing (reconciled in Gate 0, ASSUMPTIONS §3):
#   charges  -> importer.CHARGES_MAPPING / COLUMN_ALIASES
#   statements -> statements.py raw names
# Token lists are matched against header cells on an alphanumeric-only basis
# (see _norm_alnum), so "Date of Service", "DateOfService", "DATE OF SERVICE"
# all collapse to the same key.
#
# Per family:
#   prefix        - filename prefix used to route the file
#   tokens        - minimal header tokens that identify the report
#   dos_tokens    - ordered candidates for THE date-of-service column (first
#                   match wins). Explicit per family so a posting/appointment
#                   date can never be mistaken for DOS (defect: _dos_column).
#   posting_tokens- ordered candidates for a SEPARATE posting/payment date
#                   (statements only); never conflated with DOS.
#   code_tokens   - columns that are CPT/HCPCS/transaction codes: opaque strings,
#                   never numeric-coerced (defect: CPT classified as int).
REPORT_FAMILIES: Dict[str, Dict[str, object]] = {
    "charges": {
        "prefix": "ChargesHistoryDetailProviderPatientCode",
        "tokens": ("provider", "date of service", "code"),
        "dos_tokens": ("DateOfService", "Date of Service", "ServiceDate", "DOS"),
        "posting_tokens": (),
        "code_tokens": ("CPTCode", "CPT Code", "CPT", "ProcedureCode",
                        "TransactionCode", "Transaction Code", "Billed CPT",
                        "HCPCS"),
        "expected_grain": (
            "one row per charge line (patient x provider x DOS x CPT x modifier)"
        ),
    },
    "appointments": {
        "prefix": "AppointmentsPatientInfoByProviderThenDayThenFacility",
        "tokens": ("provider", "date", "status"),
        "dos_tokens": ("Appointment Date", "AppointmentDate", "DateOfService",
                       "Date of Service", "DOS"),
        "posting_tokens": (),
        "code_tokens": (),
        "expected_grain": "one row per scheduled appointment",
    },
    "statements": {
        "prefix": "PatientStatement",
        "tokens": ("date of service", "insurancepayments", "patientpayments"),
        # DOS and posting date are DIFFERENT columns and must never be confused.
        "dos_tokens": ("ChargeDateOfService", "Date of Service", "DateOfService",
                       "DOS"),
        "posting_tokens": ("StatementDate", "Statement Date", "Posting Date",
                           "PostingDate", "Payment Date", "PaymentDate"),
        "code_tokens": ("CPTCode", "CPT", "TransactionCode"),
        "expected_grain": "one row per statement grouping line (GroupingLevel3)",
    },
    "documentation": {
        "prefix": "AppointmentDocumentation",
        "tokens": ("provider", "date of service", "status"),
        "dos_tokens": ("DateOfService", "Date of Service", "DOS"),
        "posting_tokens": (),
        "code_tokens": (),
        "expected_grain": "one row per clinical note / encounter",
    },
}


def _norm(s) -> str:
    """Normalize a cell for matching: lowercase, collapse whitespace."""
    return " ".join(str(s).strip().lower().split())


def _norm_alnum(s) -> str:
    """Alphanumeric-only lowercase key, mirroring the engine's _norm_header.

    'Date of Service', 'DateOfService', 'DATE OF SERVICE' all -> 'dateofservice'.
    Used for header/token/column matching so spelling/spacing variants agree.
    """
    return "".join(c for c in str(s).lower() if c.isalnum())


def classify_filename(filename: str) -> Optional[str]:
    """Return the report-family key for a filename, or None if unrecognized."""
    base = os.path.basename(filename)
    for family, spec in REPORT_FAMILIES.items():
        if base.startswith(str(spec["prefix"])):
            return family
    return None


def detect_header_row(
    rows: Sequence[Sequence[object]],
    expected_tokens: Sequence[str],
    min_match: int = 2,
) -> int:
    """Index of the first row containing >= min_match expected tokens.

    This absorbs the preamble quirk without a hardcoded skip. Raises
    HeaderNotFoundError if no row qualifies.
    """
    wanted = [_norm_alnum(t) for t in expected_tokens]
    for i, row in enumerate(rows):
        cells = [_norm_alnum(c) for c in row]
        matches = sum(1 for t in wanted if any(t and t in c for c in cells))
        if matches >= min_match:
            return i
    raise HeaderNotFoundError(
        f"no header row matched >= {min_match} of tokens {tuple(expected_tokens)!r}"
    )


def _rows_from_csv(path: str) -> List[List[str]]:
    # utf-8-sig strips a BOM if present; Valant CSVs sometimes carry one.
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return [list(r) for r in csv.reader(f)]


def _rows_from_xlsx(path: str) -> List[List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover - environment dependent
        raise LoaderError(
            "openpyxl is required to read .xlsx exports; install it or export CSV"
        ) from e
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in row])
    finally:
        wb.close()
    return rows


def read_raw_rows(path: str) -> List[List[str]]:
    """Read every row of a raw export as lists of string cells (no parsing yet)."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".txt", ".tsv"):
        return _rows_from_csv(path)
    if ext in (".xlsx", ".xlsm"):
        return _rows_from_xlsx(path)
    raise LoaderError(f"unsupported export extension: {ext!r} ({path})")


def _assert_clean_header(header: Sequence[str], path: str) -> None:
    blanks = [i for i, h in enumerate(header) if _norm(h) == ""]
    if blanks:
        raise MissingColumnsError(
            f"blank header column(s) at index {blanks} in {os.path.basename(path)}"
        )
    seen: Dict[str, int] = {}
    dupes = set()
    for h in header:
        key = _norm(h)
        seen[key] = seen.get(key, 0) + 1
        if seen[key] > 1:
            dupes.add(h)
    if dupes:
        raise MissingColumnsError(
            f"duplicate header column(s) {sorted(dupes)} in {os.path.basename(path)}"
        )


def assert_columns(header: Sequence[str], required: Sequence[str]) -> bool:
    """Fail loudly if any required column (normalized substring match) is absent."""
    norm_header = [_norm_alnum(h) for h in header]
    missing = [
        req for req in required
        if not any(_norm_alnum(req) in h for h in norm_header)
    ]
    if missing:
        raise MissingColumnsError(f"missing required column(s): {list(missing)}")
    return True


def _match_token(header: Sequence[str], token: str) -> Optional[str]:
    """Return the first header cell whose alnum key contains the token's, else None."""
    t = _norm_alnum(token)
    if not t:
        return None
    for h in header:
        if t in _norm_alnum(h):
            return h
    return None


def resolve_dos_column(family: str, header: Sequence[str]) -> str:
    """The date-of-service column for a family, by explicit per-family tokens.

    Never falls back to a bare "date" match, so a posting or appointment date can
    never masquerade as DOS. Raises MissingColumnsError if none is present.
    """
    spec = REPORT_FAMILIES.get(family)
    if spec is None:
        raise LoaderError(f"unknown report family {family!r}")
    for token in spec.get("dos_tokens", ()):  # type: ignore[union-attr]
        col = _match_token(header, token)
        if col is not None:
            return col
    raise MissingColumnsError(
        f"no date-of-service column found for family {family!r} "
        f"(looked for {tuple(spec.get('dos_tokens', ()))!r})"
    )


def resolve_posting_column(family: str, header: Sequence[str]) -> Optional[str]:
    """The SEPARATE posting/payment-date column (statements), or None if absent.

    Returns None rather than raising: the posting date is required only for the
    Phase C maturity model, so Phase A profiling records its presence/absence.
    It is resolved from a disjoint token set so it can never equal the DOS column.
    """
    spec = REPORT_FAMILIES.get(family)
    if spec is None:
        raise LoaderError(f"unknown report family {family!r}")
    for token in spec.get("posting_tokens", ()):  # type: ignore[union-attr]
        col = _match_token(header, token)
        if col is not None:
            return col
    return None


def is_code_column(family: str, column_name: str) -> bool:
    """True if a column holds CPT/HCPCS/transaction codes for this family.

    Such columns are opaque strings and must never be numeric-coerced (a code
    like '0362T' or a leading-zero HCPCS must survive). Matched by name against
    the family's code_tokens.
    """
    spec = REPORT_FAMILIES.get(family)
    if spec is None:
        return False
    key = _norm_alnum(column_name)
    for token in spec.get("code_tokens", ()):  # type: ignore[union-attr]
        t = _norm_alnum(token)
        if t and t in key:
            return True
    return False


def load_report(
    path: str,
    expected_tokens: Optional[Sequence[str]] = None,
    family: Optional[str] = None,
) -> Tuple[List[str], List[Dict[str, str]]]:
    """Load a Valant export into (header, records).

    - Detects the header row programmatically (preamble-safe).
    - `records` is a list of dict(header_cell -> raw string cell). Values are left
      as raw strings so money/date parsing stays explicit downstream and no
      silent type coercion occurs.

    Raises HeaderNotFoundError / MissingColumnsError / LoaderError loudly; it
    never guesses past a malformed file.
    """
    if family is None:
        family = classify_filename(path)
    if expected_tokens is None:
        if family is None:
            raise LoaderError(
                f"cannot route unrecognized file {os.path.basename(path)!r}; "
                "pass expected_tokens explicitly"
            )
        expected_tokens = REPORT_FAMILIES[family]["tokens"]  # type: ignore[assignment]

    rows = read_raw_rows(path)
    if not rows:
        raise LoaderError(f"empty file: {path}")

    header_idx = detect_header_row(rows, expected_tokens)
    header = [str(c).strip() for c in rows[header_idx]]
    _assert_clean_header(header, path)

    records: List[Dict[str, str]] = []
    for raw in rows[header_idx + 1:]:
        if all(_norm(c) == "" for c in raw):
            continue  # skip fully blank trailing rows
        cells = list(raw)[: len(header)]
        cells += [""] * (len(header) - len(cells))  # pad short rows
        records.append({h: str(c) for h, c in zip(header, cells)})
    return header, records
